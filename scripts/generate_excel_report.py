import openpyxl
import json
import os
import configparser
from datetime import datetime

def main():
    config = configparser.ConfigParser()
    config.read("/home/ubuntu/daily_infrastructure_report/config/config.ini")
    data_dir = config["PATHS"]["data_dir"]
    report_dir = config["PATHS"]["report_dir"]

    # Load data
    try:
        with open(os.path.join(data_dir, "running_servers.json"), "r") as f:
            running_servers_data = json.load(f)
        with open(os.path.join(data_dir, "disk_usage.json"), "r") as f:
            disk_usage_data = json.load(f)
        with open(os.path.join(data_dir, "user_logins.json"), "r") as f:
            user_logins_data = json.load(f)
        with open(os.path.join(data_dir, "system_uptime.json"), "r") as f:
            system_uptime_data = json.load(f)
    except FileNotFoundError as e:
        print(f"Error loading data: {e}. Please ensure data collection scripts have been run.")
        return

    # Create a new Excel workbook and add sheets
    workbook = openpyxl.Workbook()

    # Running Servers Sheet
    ws_servers = workbook.active
    ws_servers.title = "Running Servers"
    ws_servers.append(["Metric", "Value"])
    ws_servers.append(["Total Running Servers", running_servers_data["count"]])
    ws_servers.append(["Running Server Hostnames"])
    for server in running_servers_data["servers"]:
        ws_servers.append([server])

    # Disk Usage Sheet
    ws_disk = workbook.create_sheet("Disk Usage")
    ws_disk.append(["Host", "Filesystem", "Size", "Used", "Avail", "Use%", "Mounted On"])
    for host, disks in disk_usage_data.items():
        for disk in disks:
            ws_disk.append([
                host,
                disk["filesystem"],
                disk["size"],
                disk["used"],
                disk["avail"],
                disk["use_percent"],
                disk["mounted_on"]
            ])

    # User Logins Sheet
    ws_logins = workbook.create_sheet("User Logins")
    ws_logins.append(["Host", "Timestamp", "Username", "Source IP", "Status"])
    for host, logins in user_logins_data.items():
        for login in logins:
            ws_logins.append([
                host,
                login["timestamp"],
                login["username"],
                login["source_ip"],
                login["status"]
            ])

    # System Uptime Sheet
    ws_uptime = workbook.create_sheet("System Uptime")
    ws_uptime.append(["Host", "Uptime"])
    for host, uptime in system_uptime_data.items():
        ws_uptime.append([host, uptime])

    if not os.path.exists(report_dir):
        os.makedirs(report_dir)

    output_excel_path = os.path.join(report_dir, f"infrastructure_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    workbook.save(output_excel_path)
    print(f"Excel report generated: {output_excel_path}")

if __name__ == "__main__":
    main()


